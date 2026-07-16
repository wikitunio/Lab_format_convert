import streamlit as st
import openpyxl
from io import BytesIO
import datetime  # Added to handle exact date formatting

st.set_page_config(page_title="UREA Shift Data Sync", layout="centered")

# --- CSS STYLING FOR THE FOOTER ---
st.markdown("""
    <style>
    .footer {
        position: fixed;
        left: 0;
        bottom: 0;
        width: 100%;
        background-color: transparent;
        color: gray;
        text-align: center;
        padding: 15px;
        font-size: 14px;
        border-top: 1px solid #eaeaea;
    }
    .footer a {
        color: #0072b1; /* LinkedIn Blue */
        text-decoration: none;
        font-weight: bold;
    }
    .footer a:hover {
        text-decoration: underline;
    }
    </style>
""", unsafe_allow_html=True)

st.title("Daily UREA Plant Data Sync")
st.markdown("Upload your daily analytical report to auto-populate the DAR template.")

# Hardcoded mapping dictionary
CELL_MAPPING = {
    'A1': 'A2', 'C7': 'B5', 'C8': 'C5', 'C9': 'D5', 'D7': 'E5', 'D8': 'F5', 'D9': 'G5',
    'E7': 'H5', 'E8': 'I5', 'E9': 'J5', 'F7': 'K5', 'F8': 'L5', 'F9': 'M5', 'G7': 'N5',
    'G8': 'O5', 'H7': 'Q5', 'H8': 'R5', 'I7': 'N8', 'I8': 'O8', 'I9': 'P8', 'J7': 'B8',
    'J8': 'C8', 'K7': 'E8', 'K8': 'F8', 'L7': 'H8', 'L8': 'I8', 'L9': 'J8', 'M7': 'K8',
    'N7': 'Q8', 'C18': 'C11', 'C19': 'E11', 'C20': 'D11', 'C21': 'F11', 'C22': 'G11',
    'C23': 'H11', 'C24': 'I11', 'C25': 'J11', 'C26': 'K11', 'C27': 'L11', 'C28': 'M11',
    'C29': 'N11', 'C30': 'O11', 'C31': 'P11', 'C32': 'B11', 'D18': 'C12', 'D19': 'E12',
    'D20': 'D12', 'D21': 'F12', 'D22': 'G12', 'D23': 'H12', 'D24': 'I12', 'D25': 'J12',
    'D26': 'K12', 'D27': 'L12', 'D28': 'M12', 'D29': 'N12', 'D30': 'O12', 'D31': 'P12',
    'D32': 'B12', 'E18': 'C13', 'E19': 'E13', 'E20': 'D13', 'E21': 'F13', 'E22': 'G13',
    'E23': 'H13', 'E24': 'I13', 'E25': 'J13', 'E26': 'K13', 'E27': 'L13', 'E28': 'M13',
    'E29': 'N13', 'E30': 'O13', 'E31': 'P13', 'E32': 'B13', 'R17': 'B16', 'R18': 'B17',
    'R19': 'F16', 'R20': 'E16', 'R21': 'D16'
}

# Cache the template file in memory to avoid repeated disk reads
@st.cache_data
def load_template():
    with open("DAR 27-06-2026.xlsx", "rb") as f:
        return f.read()

source_file = st.file_uploader("Upload Daily Analysis Report DAR (must contain 'UREA' sheet)", type=["xlsx", "xls"])

if source_file is not None:
    # Provide visual feedback during the blocking file read operation
    with st.spinner("Processing data and generating report..."):
        try:
            # Load the uploaded source file
            source_wb = openpyxl.load_workbook(source_file, data_only=True)
            
            if "UREA" in source_wb.sheetnames:
                source_sheet = source_wb["UREA"]
                
                # Extract raw date from A1
                raw_date = source_sheet['A1'].value
                
                # Format the date safely for the filename
                if isinstance(raw_date, datetime.datetime):
                    # If Excel stored it as a true date, format it nicely (e.g., 27-06-2026)
                    safe_date = raw_date.strftime("%d-%m-%Y")
                elif raw_date:
                    # If Excel stored it as text, clean out characters that break filenames
                    safe_date = str(raw_date).replace("/", "-").replace(":", "-").split()[0]
                else:
                    # Fallback if A1 is totally empty
                    safe_date = "Updated"
                
                # Load the cached DAR template from fast memory instead of disk
                target_wb = openpyxl.load_workbook(BytesIO(load_template()))
                target_sheet = target_wb["Sheet1"] 
                
                # Transfer data quickly
                for source_cell, target_cell in CELL_MAPPING.items():
                    target_sheet[target_cell].value = source_sheet[source_cell].value
                
                # Save to memory buffer
                output = BytesIO()
                target_wb.save(output)
                output.seek(0)
                
                # Explicitly close workbooks to prevent memory leaks
                source_wb.close()
                target_wb.close()
                
                st.success(f"✅ Report for **{safe_date}** generated successfully!")
                    
                # Provide the download button with the dynamic filename
                st.download_button(
                    label=f"📥 Download DAR Report ({safe_date})",
                    data=output,
                    file_name=f"DAR {safe_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            else:
                st.error("The uploaded file does not contain a sheet named 'UREA'. Please verify the file.")
                source_wb.close()
                
        except Exception as e:
            st.error(f"An error occurred while processing the files: {e}")

# --- FOOTER SECTION ---
st.markdown(f"""
    <div class="footer">
        Developed by <a href="https://www.linkedin.com/in/wikitunio" target="_blank">Waqar Ahmed Tunio</a> <br>
        Email: <a href="mailto:ahmed.waqar@pafl.com.pk">ahmed.waqar@pafl.com.pk</a>
    </div>
    """, unsafe_allow_html=True)
